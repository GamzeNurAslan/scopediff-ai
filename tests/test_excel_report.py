from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.database.models import (
    AnalysisRun,
    DefectRanking,
    RequirementChange,
)
from backend.app.reports.excel_report import (
    ExcelReportGenerator,
)


def create_analysis() -> AnalysisRun:
    analysis = AnalysisRun(
        id=7,
        analysis_name="v1 to v2 analysis",
        source_version="v1",
        target_version="v2",
        created_at=datetime(
            2026,
            8,
            7,
            12,
            30,
        ),
    )

    analysis.requirement_changes.append(
        RequirementChange(
            id=1,
            old_requirement_id="REQ-001",
            new_requirement_id="REQ-001",
            change_type="numeric_change",
            risk_score=65.0,
            risk_level="high",
            confidence=0.91,
            explanation=(
                "Sayısal sınır değişti."
            ),
        )
    )

    analysis.requirement_changes.append(
        RequirementChange(
            id=2,
            old_requirement_id="REQ-002",
            new_requirement_id="REQ-002",
            change_type="negation_change",
            risk_score=82.0,
            risk_level="critical",
            confidence=0.87,
            explanation=(
                "Negation değişikliği tespit edildi."
            ),
        )
    )

    analysis.defect_rankings.append(
        DefectRanking(
            id=1,
            defect_id="DEF-001",
            defect_text=(
                "Port kontrolü sürekli "
                "tekrar ediyor."
            ),
            change_id="TR-001-2",
            relevance_score=0.88,
            rank_position=1,
            reason=(
                "Kesin kök neden değildir; "
                "incelenmesi gereken aday "
                "değişikliklerden biridir."
            ),
        )
    )

    return analysis


def test_report_file_is_created(
    tmp_path: Path,
) -> None:
    generator = ExcelReportGenerator()

    output_path = (
        tmp_path
        / "ScopeDiff_Report.xlsx"
    )

    result = generator.generate(
        create_analysis(),
        output_path,
    )

    assert result == output_path
    assert output_path.exists()


def test_report_contains_expected_sheets(
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "report.xlsx"
    )

    ExcelReportGenerator().generate(
        create_analysis(),
        output_path,
    )

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    assert workbook.sheetnames == [
        "Summary",
        "Requirement Changes",
        "Defect Rankings",
    ]


def test_summary_contains_analysis_metadata(
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "report.xlsx"
    )

    ExcelReportGenerator().generate(
        create_analysis(),
        output_path,
    )

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    sheet = workbook["Summary"]

    assert sheet["B3"].value == 7

    assert (
        sheet["B4"].value
        == "v1 to v2 analysis"
    )

    assert sheet["B5"].value == "v1"
    assert sheet["B6"].value == "v2"


def test_summary_uses_formulas_for_metrics(
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "report.xlsx"
    )

    ExcelReportGenerator().generate(
        create_analysis(),
        output_path,
    )

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    sheet = workbook["Summary"]

    assert (
        sheet["B10"].value
        == (
            "=COUNTA("
            "'Requirement Changes'!A:A"
            ")-1"
        )
    )

    assert (
        sheet["B11"].value
        == (
            "=COUNTA("
            "'Defect Rankings'!A:A"
            ")-1"
        )
    )

    assert str(
        sheet["B17"].value
    ).startswith(
        "=COUNTIF("
    )


def test_requirement_changes_are_written(
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "report.xlsx"
    )

    ExcelReportGenerator().generate(
        create_analysis(),
        output_path,
    )

    workbook = load_workbook(
        output_path,
    )

    sheet = workbook[
        "Requirement Changes"
    ]

    assert sheet.max_row == 3

    assert sheet["B2"].value == "REQ-001"

    assert (
        sheet["D2"].value
        == "numeric_change"
    )

    assert sheet["E2"].value == 65.0

    assert (
        sheet["F3"].value
        == "critical"
    )


def test_defect_rankings_are_written(
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "report.xlsx"
    )

    ExcelReportGenerator().generate(
        create_analysis(),
        output_path,
    )

    workbook = load_workbook(
        output_path,
    )

    sheet = workbook[
        "Defect Rankings"
    ]

    assert sheet.max_row == 2

    assert (
        sheet["B2"].value
        == "DEF-001"
    )

    assert (
        sheet["D2"].value
        == "TR-001-2"
    )

    assert sheet["E2"].value == 0.88

    assert sheet["F2"].value == 1


def test_invalid_extension_is_rejected(
    tmp_path: Path,
) -> None:
    generator = ExcelReportGenerator()

    with pytest.raises(
        ValueError
    ) as exception:
        generator.generate(
            create_analysis(),
            tmp_path / "report.csv",
        )

    assert ".xlsx" in str(
        exception.value
    )