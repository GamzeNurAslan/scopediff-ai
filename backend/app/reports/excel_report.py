from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)

from openpyxl.worksheet.worksheet import (
    Worksheet,
)

from backend.app.database.models import (
    AnalysisRun,
)


class ExcelReportGenerator:
    """
    ScopeDiff analiz sonuçlarını
    çok dilli Excel raporuna dönüştürür.

    Requirement ve defect metinlerinin
    kendisi çevrilmez.

    Yalnızca rapor yapısı, kolonlar,
    risk etiketleri ve bilinen sistem
    change type değerleri yerelleştirilir.
    """

    NAVY = "14213D"
    DARK_NAVY = "0B132B"
    ORANGE = "F59E0B"

    WHITE = "FFFFFF"

    STATIC_GRAY = "6B7280"


    RISK_COLORS = {
        "low":
            "DCFCE7",

        "medium":
            "FEF3C7",

        "high":
            "FED7AA",

        "critical":
            "FECACA",
    }


    SUPPORTED_LANGUAGES = {
        "tr",
        "en",
        "de",
        "fr",
        "es",
    }


    TEXT = {
        "en": {
            "sheet.summary":
                "Summary",

            "sheet.changes":
                "Requirement Changes",

            "sheet.defects":
                "Defect Rankings",

            "title":
                "ScopeDiff AI - Analysis Report",

            "analysis_id":
                "Analysis ID",

            "analysis_name":
                "Analysis Name",

            "source_version":
                "Source Version",

            "target_version":
                "Target Version",

            "created_at":
                "Created At",

            "analysis_metrics":
                "Analysis Metrics",

            "total_changes":
                "Total Requirement Changes",

            "total_defects":
                "Total Defect Rankings",

            "average_risk":
                "Average Risk Score",

            "average_relevance":
                "Average Relevance Score",

            "risk_distribution":
                "Risk Distribution",

            "old_requirement_id":
                "Old Requirement ID",

            "new_requirement_id":
                "New Requirement ID",

            "change_type":
                "Change Type",

            "risk_score":
                "Risk Score",

            "risk_level":
                "Risk Level",

            "confidence":
                "Confidence",

            "explanation":
                "Explanation",

            "defect_id":
                "Defect ID",

            "defect_text":
                "Defect Text",

            "change_id":
                "Change ID",

            "relevance":
                "Relevance Score",

            "rank":
                "Rank Position",

            "reason":
                "Reason",

            "risk.low":
                "Low",

            "risk.medium":
                "Medium",

            "risk.high":
                "High",

            "risk.critical":
                "Critical",

            "change.unchanged":
                "Unchanged",

            "change.paraphrased":
                "Paraphrased",

            "change.changed":
                "Changed",

            "change.added":
                "Added",

            "change.removed":
                "Removed",
        },


        "tr": {
            "sheet.summary":
                "Özet",

            "sheet.changes":
                "Gereksinim Değişiklikleri",

            "sheet.defects":
                "Defect Sıralamaları",

            "title":
                "ScopeDiff AI - Analiz Raporu",

            "analysis_id":
                "Analiz ID",

            "analysis_name":
                "Analiz Adı",

            "source_version":
                "Kaynak Versiyon",

            "target_version":
                "Hedef Versiyon",

            "created_at":
                "Oluşturulma Tarihi",

            "analysis_metrics":
                "Analiz Metrikleri",

            "total_changes":
                "Toplam Gereksinim Değişikliği",

            "total_defects":
                "Toplam Defect Aday Kaydı",

            "average_risk":
                "Ortalama Risk Skoru",

            "average_relevance":
                "Ortalama İlgililik Skoru",

            "risk_distribution":
                "Risk Dağılımı",

            "old_requirement_id":
                "Eski Gereksinim ID",

            "new_requirement_id":
                "Yeni Gereksinim ID",

            "change_type":
                "Değişiklik Türü",

            "risk_score":
                "Risk Skoru",

            "risk_level":
                "Risk Seviyesi",

            "confidence":
                "Confidence",

            "explanation":
                "Açıklama",

            "defect_id":
                "Defect ID",

            "defect_text":
                "Defect Metni",

            "change_id":
                "Değişiklik ID",

            "relevance":
                "İlgililik Skoru",

            "rank":
                "Sıralama",

            "reason":
                "Gerekçe",

            "risk.low":
                "Düşük",

            "risk.medium":
                "Orta",

            "risk.high":
                "Yüksek",

            "risk.critical":
                "Kritik",

            "change.unchanged":
                "Değişmedi",

            "change.paraphrased":
                "Yeniden İfade",

            "change.changed":
                "Değiştirildi",

            "change.added":
                "Eklendi",

            "change.removed":
                "Kaldırıldı",
        },


        "de": {
            "sheet.summary":
                "Übersicht",

            "sheet.changes":
                "Anforderungsänderungen",

            "sheet.defects":
                "Defekt-Rangliste",

            "title":
                "ScopeDiff AI - Analysebericht",

            "analysis_id":
                "Analyse-ID",

            "analysis_name":
                "Analysename",

            "source_version":
                "Quellversion",

            "target_version":
                "Zielversion",

            "created_at":
                "Erstellt am",

            "analysis_metrics":
                "Analysemetriken",

            "total_changes":
                "Anzahl Anforderungsänderungen",

            "total_defects":
                "Anzahl Defekt-Ranglisten",

            "average_risk":
                "Durchschnittlicher Risikowert",

            "average_relevance":
                "Durchschnittliche Relevanz",

            "risk_distribution":
                "Risikoverteilung",

            "old_requirement_id":
                "Alte Anforderungs-ID",

            "new_requirement_id":
                "Neue Anforderungs-ID",

            "change_type":
                "Änderungstyp",

            "risk_score":
                "Risikowert",

            "risk_level":
                "Risikostufe",

            "confidence":
                "Konfidenz",

            "explanation":
                "Erklärung",

            "defect_id":
                "Defekt-ID",

            "defect_text":
                "Defekttext",

            "change_id":
                "Änderungs-ID",

            "relevance":
                "Relevanzwert",

            "rank":
                "Rangposition",

            "reason":
                "Begründung",

            "risk.low":
                "Niedrig",

            "risk.medium":
                "Mittel",

            "risk.high":
                "Hoch",

            "risk.critical":
                "Kritisch",

            "change.unchanged":
                "Unverändert",

            "change.paraphrased":
                "Umformuliert",

            "change.changed":
                "Geändert",

            "change.added":
                "Hinzugefügt",

            "change.removed":
                "Entfernt",
        },


        "fr": {
            "sheet.summary":
                "Résumé",

            "sheet.changes":
                "Modifications Exigences",

            "sheet.defects":
                "Classement Défauts",

            "title":
                "ScopeDiff AI - Rapport d’analyse",

            "analysis_id":
                "ID d’analyse",

            "analysis_name":
                "Nom de l’analyse",

            "source_version":
                "Version source",

            "target_version":
                "Version cible",

            "created_at":
                "Créé le",

            "analysis_metrics":
                "Métriques d’analyse",

            "total_changes":
                "Total des modifications",

            "total_defects":
                "Total des classements de défauts",

            "average_risk":
                "Score de risque moyen",

            "average_relevance":
                "Score de pertinence moyen",

            "risk_distribution":
                "Répartition des risques",

            "old_requirement_id":
                "Ancien ID exigence",

            "new_requirement_id":
                "Nouvel ID exigence",

            "change_type":
                "Type de changement",

            "risk_score":
                "Score de risque",

            "risk_level":
                "Niveau de risque",

            "confidence":
                "Confiance",

            "explanation":
                "Explication",

            "defect_id":
                "ID défaut",

            "defect_text":
                "Texte du défaut",

            "change_id":
                "ID changement",

            "relevance":
                "Score de pertinence",

            "rank":
                "Position",

            "reason":
                "Raison",

            "risk.low":
                "Faible",

            "risk.medium":
                "Moyen",

            "risk.high":
                "Élevé",

            "risk.critical":
                "Critique",

            "change.unchanged":
                "Inchangé",

            "change.paraphrased":
                "Reformulé",

            "change.changed":
                "Modifié",

            "change.added":
                "Ajouté",

            "change.removed":
                "Supprimé",
        },


        "es": {
            "sheet.summary":
                "Resumen",

            "sheet.changes":
                "Cambios Requisitos",

            "sheet.defects":
                "Ranking Defectos",

            "title":
                "ScopeDiff AI - Informe de análisis",

            "analysis_id":
                "ID del análisis",

            "analysis_name":
                "Nombre del análisis",

            "source_version":
                "Versión de origen",

            "target_version":
                "Versión de destino",

            "created_at":
                "Fecha de creación",

            "analysis_metrics":
                "Métricas del análisis",

            "total_changes":
                "Total de cambios de requisitos",

            "total_defects":
                "Total de rankings de defectos",

            "average_risk":
                "Puntuación media de riesgo",

            "average_relevance":
                "Relevancia media",

            "risk_distribution":
                "Distribución de riesgos",

            "old_requirement_id":
                "ID requisito anterior",

            "new_requirement_id":
                "ID requisito nuevo",

            "change_type":
                "Tipo de cambio",

            "risk_score":
                "Puntuación de riesgo",

            "risk_level":
                "Nivel de riesgo",

            "confidence":
                "Confianza",

            "explanation":
                "Explicación",

            "defect_id":
                "ID defecto",

            "defect_text":
                "Texto del defecto",

            "change_id":
                "ID cambio",

            "relevance":
                "Puntuación de relevancia",

            "rank":
                "Posición",

            "reason":
                "Motivo",

            "risk.low":
                "Bajo",

            "risk.medium":
                "Medio",

            "risk.high":
                "Alto",

            "risk.critical":
                "Crítico",

            "change.unchanged":
                "Sin cambios",

            "change.paraphrased":
                "Parafraseado",

            "change.changed":
                "Modificado",

            "change.added":
                "Añadido",

            "change.removed":
                "Eliminado",
        },
    }


    def _language(
        self,
        language: str,
    ) -> str:
        normalized = (
            language
            .strip()
            .lower()
        )

        if (
            normalized
            in self.SUPPORTED_LANGUAGES
        ):
            return normalized

        return "en"


    def _t(
        self,
        language: str,
        key: str,
    ) -> str:
        current = (
            self.TEXT[
                language
            ].get(
                key
            )
        )

        if current is not None:
            return current

        return (
            self.TEXT[
                "en"
            ].get(
                key,
                key,
            )
        )


    def _risk_label(
        self,
        risk: object,
        language: str,
        preserve_code: bool = False,
    ) -> str:
        normalized = str(
            risk or ""
        ).strip().lower()

        if preserve_code:
            return normalized

        return self._t(
            language,
            f"risk.{normalized}",
        )


    def _change_type_label(
        self,
        change_type: object,
        language: str,
        preserve_code: bool = False,
    ) -> str:
        normalized = (
            str(
                change_type
                or ""
            )
            .strip()
            .lower()
        )

        if preserve_code:
            return normalized

        translated = (
            self.TEXT[
                language
            ].get(
                f"change.{normalized}"
            )
        )

        if translated:
            return translated

        return (
            normalized
            .replace(
                "_",
                " ",
            )
            .title()
        )


    @staticmethod
    def _sheet_formula_name(
        value: str,
    ) -> str:
        return value.replace(
            "'",
            "''",
        )


    def generate(
        self,
        analysis: AnalysisRun,
        output_path: str | Path,
        language: str | None = None,
    ) -> Path:
        preserve_change_codes = language is None
        language = self._language(
            language or "en"
        )

        path = Path(
            output_path
        )

        if (
            path.suffix.lower()
            != ".xlsx"
        ):
            raise ValueError(
                "Excel report extension "
                "must be .xlsx."
            )

        path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()

        summary_title = self._t(
            language,
            "sheet.summary",
        )

        changes_title = self._t(
            language,
            "sheet.changes",
        )

        defects_title = self._t(
            language,
            "sheet.defects",
        )

        summary_sheet = (
            workbook.active
        )

        summary_sheet.title = (
            summary_title
        )

        changes_sheet = (
            workbook.create_sheet(
                changes_title
            )
        )

        rankings_sheet = (
            workbook.create_sheet(
                defects_title
            )
        )

        self._build_summary_sheet(
            summary_sheet,
            analysis,
            language,
            changes_title,
            defects_title,
        )

        self._build_requirement_changes_sheet(
            changes_sheet,
            analysis,
            language,
            preserve_change_codes,
        )

        self._build_defect_rankings_sheet(
            rankings_sheet,
            analysis,
            language,
        )

        workbook.save(
            path
        )

        return path


    def _build_summary_sheet(
        self,
        sheet: Worksheet,
        analysis: AnalysisRun,
        language: str,
        changes_sheet_title: str,
        defects_sheet_title: str,
    ) -> None:
        sheet.sheet_view.showGridLines = (
            False
        )

        sheet.sheet_properties.tabColor = (
            self.ORANGE
        )

        sheet.merge_cells(
            "A1:D1"
        )

        title_cell = (
            sheet["A1"]
        )

        title_cell.value = (
            self._t(
                language,
                "title",
            )
        )

        title_cell.font = Font(
            bold=True,
            size=18,
            color=self.WHITE,
        )

        title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=self.DARK_NAVY,
        )

        title_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        sheet.row_dimensions[
            1
        ].height = 30


        metadata = [
            (
                self._t(
                    language,
                    "analysis_id",
                ),
                (
                    analysis.id
                    if analysis.id
                    is not None
                    else "-"
                ),
            ),
            (
                self._t(
                    language,
                    "analysis_name",
                ),
                analysis.analysis_name,
            ),
            (
                self._t(
                    language,
                    "source_version",
                ),
                (
                    analysis
                    .source_version
                    or "-"
                ),
            ),
            (
                self._t(
                    language,
                    "target_version",
                ),
                (
                    analysis
                    .target_version
                    or "-"
                ),
            ),
            (
                self._t(
                    language,
                    "created_at",
                ),
                (
                    analysis
                    .created_at
                    or "-"
                ),
            ),
        ]


        for offset, (
            label,
            value,
        ) in enumerate(
            metadata,
            start=3,
        ):
            sheet.cell(
                row=offset,
                column=1,
                value=label,
            )

            sheet.cell(
                row=offset,
                column=2,
                value=value,
            )

            self._style_summary_label(
                sheet.cell(
                    row=offset,
                    column=1,
                )
            )


        sheet["A9"] = self._t(
            language,
            "analysis_metrics",
        )

        self._style_section_title(
            sheet["A9"]
        )


        changes_formula_name = (
            self._sheet_formula_name(
                changes_sheet_title
            )
        )

        defects_formula_name = (
            self._sheet_formula_name(
                defects_sheet_title
            )
        )


        metrics = [
            (
                self._t(
                    language,
                    "total_changes",
                ),
                (
                    "=COUNTA("
                    f"'{changes_formula_name}'"
                    "!A:A)-1"
                ),
            ),
            (
                self._t(
                    language,
                    "total_defects",
                ),
                (
                    "=COUNTA("
                    f"'{defects_formula_name}'"
                    "!A:A)-1"
                ),
            ),
            (
                self._t(
                    language,
                    "average_risk",
                ),
                (
                    "=IFERROR("
                    "AVERAGE("
                    f"'{changes_formula_name}'"
                    "!E:E),0)"
                ),
            ),
            (
                self._t(
                    language,
                    "average_relevance",
                ),
                (
                    "=IFERROR("
                    "AVERAGE("
                    f"'{defects_formula_name}'"
                    "!E:E),0)"
                ),
            ),
        ]


        for row, (
            label,
            formula,
        ) in enumerate(
            metrics,
            start=10,
        ):
            sheet.cell(
                row=row,
                column=1,
                value=label,
            )

            sheet.cell(
                row=row,
                column=2,
                value=formula,
            )

            self._style_summary_label(
                sheet.cell(
                    row=row,
                    column=1,
                )
            )


        sheet["B12"].number_format = (
            "0.0"
        )

        sheet["B13"].number_format = (
            "0.0%"
        )


        sheet["A16"] = self._t(
            language,
            "risk_distribution",
        )

        self._style_section_title(
            sheet["A16"]
        )


        risks = [
            "low",
            "medium",
            "high",
            "critical",
        ]


        for row, risk in enumerate(
            risks,
            start=17,
        ):
            risk_label = (
                self._risk_label(
                    risk,
                    language,
                )
            )

            sheet.cell(
                row=row,
                column=1,
                value=risk_label,
            )

            sheet.cell(
                row=row,
                column=2,
                value=(
                    "=COUNTIF("
                    f"'{changes_formula_name}'"
                    "!F:F,"
                    f'"{risk_label}"'
                    ")"
                ),
            )

            risk_cell = (
                sheet.cell(
                    row=row,
                    column=1,
                )
            )

            risk_cell.fill = (
                PatternFill(
                    fill_type="solid",
                    fgColor=(
                        self.RISK_COLORS[
                            risk
                        ]
                    ),
                )
            )

            risk_cell.font = Font(
                bold=True
            )


        if (
            analysis.created_at
            is not None
        ):
            sheet["B7"].number_format = (
                "yyyy-mm-dd hh:mm"
            )


        sheet.column_dimensions[
            "A"
        ].width = 34

        sheet.column_dimensions[
            "B"
        ].width = 30

        sheet.column_dimensions[
            "C"
        ].width = 18

        sheet.column_dimensions[
            "D"
        ].width = 18


    def _build_requirement_changes_sheet(
        self,
        sheet: Worksheet,
        analysis: AnalysisRun,
        language: str,
        preserve_change_codes: bool = False,
    ) -> None:
        headers = [
            "ID",

            self._t(
                language,
                "old_requirement_id",
            ),

            self._t(
                language,
                "new_requirement_id",
            ),

            self._t(
                language,
                "change_type",
            ),

            self._t(
                language,
                "risk_score",
            ),

            self._t(
                language,
                "risk_level",
            ),

            self._t(
                language,
                "confidence",
            ),

            self._t(
                language,
                "explanation",
            ),
        ]


        self._write_table_header(
            sheet,
            headers,
        )


        for row_index, change in enumerate(
            analysis.requirement_changes,
            start=2,
        ):
            risk_level = str(
                change.risk_level
            ).strip().lower()

            values = [
                change.id,

                change.old_requirement_id,

                change.new_requirement_id,

                self._change_type_label(
                    change.change_type,
                    language,
                    preserve_change_codes,
                ),

                change.risk_score,

                self._risk_label(
                    risk_level,
                    language,
                    preserve_change_codes,
                ),

                change.confidence,

                change.explanation,
            ]


            for (
                column_index,
                value,
            ) in enumerate(
                values,
                start=1,
            ):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )


            sheet.cell(
                row=row_index,
                column=5,
            ).number_format = (
                "0.0"
            )


            sheet.cell(
                row=row_index,
                column=7,
            ).number_format = (
                "0.0%"
            )


            if (
                risk_level
                in self.RISK_COLORS
            ):
                sheet.cell(
                    row=row_index,
                    column=6,
                ).fill = (
                    PatternFill(
                        fill_type="solid",
                        fgColor=(
                            self.RISK_COLORS[
                                risk_level
                            ]
                        ),
                    )
                )


            sheet.cell(
                row=row_index,
                column=8,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


        sheet.freeze_panes = (
            "A2"
        )

        sheet.auto_filter.ref = (
            f"A1:H{max(sheet.max_row, 1)}"
        )


        self._set_column_widths(
            sheet,
            {
                "A": 10,
                "B": 23,
                "C": 23,
                "D": 24,
                "E": 15,
                "F": 16,
                "G": 14,
                "H": 55,
            },
        )


    def _build_defect_rankings_sheet(
        self,
        sheet: Worksheet,
        analysis: AnalysisRun,
        language: str,
    ) -> None:
        headers = [
            "ID",

            self._t(
                language,
                "defect_id",
            ),

            self._t(
                language,
                "defect_text",
            ),

            self._t(
                language,
                "change_id",
            ),

            self._t(
                language,
                "relevance",
            ),

            self._t(
                language,
                "rank",
            ),

            self._t(
                language,
                "reason",
            ),
        ]


        self._write_table_header(
            sheet,
            headers,
        )


        rankings = sorted(
            analysis.defect_rankings,
            key=lambda item: (
                item.defect_id
                or "",

                item.rank_position,
            ),
        )


        for (
            row_index,
            ranking,
        ) in enumerate(
            rankings,
            start=2,
        ):
            values = [
                ranking.id,

                ranking.defect_id,

                ranking.defect_text,

                ranking.change_id,

                ranking.relevance_score,

                ranking.rank_position,

                ranking.reason,
            ]


            for (
                column_index,
                value,
            ) in enumerate(
                values,
                start=1,
            ):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )


            sheet.cell(
                row=row_index,
                column=5,
            ).number_format = (
                "0.0%"
            )


            sheet.cell(
                row=row_index,
                column=3,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


            sheet.cell(
                row=row_index,
                column=7,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


        sheet.freeze_panes = (
            "A2"
        )

        sheet.auto_filter.ref = (
            f"A1:G{max(sheet.max_row, 1)}"
        )


        self._set_column_widths(
            sheet,
            {
                "A": 10,
                "B": 16,
                "C": 48,
                "D": 18,
                "E": 20,
                "F": 18,
                "G": 60,
            },
        )


    def _write_table_header(
        self,
        sheet: Worksheet,
        headers: list[str],
    ) -> None:
        sheet.sheet_view.showGridLines = (
            False
        )

        sheet.sheet_properties.tabColor = (
            self.NAVY
        )


        for (
            column_index,
            header,
        ) in enumerate(
            headers,
            start=1,
        ):
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color=self.WHITE,
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.NAVY,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


        sheet.row_dimensions[
            1
        ].height = 24


    def _style_summary_label(
        self,
        cell,
    ) -> None:
        cell.font = Font(
            bold=True,
            color=self.STATIC_GRAY,
        )


    def _style_section_title(
        self,
        cell,
    ) -> None:
        cell.font = Font(
            bold=True,
            color=self.WHITE,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=self.NAVY,
        )


    @staticmethod
    def _set_column_widths(
        sheet: Worksheet,
        widths:
            dict[str, float],
    ) -> None:
        for (
            column,
            width,
        ) in widths.items():
            sheet.column_dimensions[
                column
            ].width = width
