from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)


def main() -> None:
    project_root = (
        Path(__file__)
        .resolve()
        .parents[2]
    )

    output_dir = (
        project_root
        / "sample_files"
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path = (
        output_dir
        / "process_tracking_sample.xlsx"
    )

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Süreç Takibi"

    headers = [
        "Süreç",
        "Portal Menüsü",
        "Bitiş Tarihi/Durum",
        "Yazılımcı",
        "Teste Verildi mi?",
        "Analist",
        "ANALİZ OK/NOK",
        "Test OK/NOK",
        "Modül",
        "Notlar",
    ]

    rows = [
        [
            "Port Tahsis geliştirmesi",
            "Diğer İşlemler > Manuel Port Tahsis",
            date(2026, 8, 20),
            "Developer A",
            "Evet",
            "Analist A",
            "OK",
            "",
            "xDSL",
            "Test süreci bekleniyor.",
        ],
        [
            "Ön Başvuru Listesi",
            "Diğer İşlemler > Ön Başvuru Listesi",
            date(2026, 8, 22),
            "Developer B",
            "",
            "Analist B",
            "OK",
            "",
            "xDSL",
            "",
        ],
        [
            "Ön Başvuru Açıklama",
            "Diğer İşlemler > Ön Başvuru Açıklama",
            date(2026, 8, 19),
            "Developer C",
            "Evet",
            "Analist C",
            "OK",
            "NOK",
            "xDSL",
            "Test sırasında yeniden incelenmeli.",
        ],
        [
            "FTTX Abone İzleme",
            "Diğer İşlemler > FTTX Abone İzleme",
            date(2026, 8, 25),
            "Developer D",
            "Evet",
            "Analist D",
            "OK",
            "OK",
            "FTTX",
            "Test tamamlandı.",
        ],
        [
            "Nakil İşlemi İptali",
            "Değişiklik İşlemleri > Nakil İşlemi İptali",
            date(2026, 8, 27),
            "Developer E",
            "",
            "Analist E",
            "NOK",
            "",
            "xDSL",
            "Analiz tarafında tekrar incelenecek.",
        ],
        [
            "Yalın DSL Kontrolü",
            "Diğer Sorgulamalar > Yalın DSL",
            "Teslime Hazır",
            "Developer F",
            "Evet",
            "Analist F",
            "OK",
            "OK",
            "DSL",
            "",
        ],
        [
            "Kurumsal Müşteri SMS",
            "Müşteri İşlemleri",
            "Teslim Edildi",
            "Developer G",
            "Evet",
            "Analist G",
            "OK",
            "OK",
            "CRM",
            "Tamamlandı.",
        ],
        [
            "Yeni Aktivasyon Akışı",
            "Aktivasyon > Yeni Başvuru",
            date(2026, 8, 30),
            "",
            "Hayır",
            "Analist H",
            "",
            "",
            "Activation",
            "Tasarım aşamasında.",
        ],
    ]

    for (
        column_index,
        header,
    ) in enumerate(
        headers,
        start=1,
    ):
        cell = sheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="172554",
        )

        cell.alignment = Alignment(
            vertical="center",
        )

    for (
        row_index,
        values,
    ) in enumerate(
        rows,
        start=2,
    ):
        for (
            column_index,
            value,
        ) in enumerate(
            values,
            start=1,
        ):
            sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    widths = {
        "A": 32,
        "B": 42,
        "C": 22,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 42,
    }

    for (
        column,
        width,
    ) in widths.items():
        sheet.column_dimensions[
            column
        ].width = width

    sheet.freeze_panes = "A2"

    workbook.save(
        output_path
    )

    print(
        "OK - sample process file created:"
    )

    print(
        output_path
    )


if __name__ == "__main__":
    main()